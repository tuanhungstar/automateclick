# File: Bot_module/Excel_Module.py

import sys
import os
import openpyxl
import pandas as pd
from typing import Optional, List, Dict, Any, Union
from datetime import datetime

# --- PyQt6 Imports ---
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QFormLayout, QLineEdit, QPushButton, QDialogButtonBox,
    QComboBox, QWidget, QGroupBox, QMessageBox, QLabel,
    QCheckBox, QFileDialog, QHBoxLayout, QRadioButton, QSpinBox, 
    QScrollArea, QFrame
)
from PyQt6.QtCore import Qt

# --- Main App Imports (Fallback for standalone testing) ---
try:
    from my_lib.shared_context import ExecutionContext
except ImportError:
    class ExecutionContext:
        def __init__(self): self.vars = {}
        def add_log(self, message: str): print(f"LOG: {message}")
        def get_variable(self, name: str, default: Any = None) -> Any: return self.vars.get(name, default)
        def set_variable(self, name: str, value: Any): self.vars[name] = value

#
# --- HELPER: Action Row for Excel Read ---
#
class _ExcelReadActionRow(QFrame):
    def __init__(self, global_variables: List[str], index: int, parent=None):
        super().__init__(parent)
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.index = index
        self.global_variables = ["-- Select Variable --"] + global_variables
        
        layout = QVBoxLayout(self)
        top_layout = QHBoxLayout()
        self.action_combo = QComboBox()
        self.action_combo.addItems(["Read Cell", "Get Last Empty Row", "Get Total Line Count", "Find First Empty Row (1-Max)"])
        self.delete_btn = QPushButton("Remove"); self.delete_btn.setStyleSheet("background-color: #ff4d4d; color: white;")
        top_layout.addWidget(QLabel(f"Action {index+1}:")); top_layout.addWidget(self.action_combo, 1); top_layout.addWidget(self.delete_btn)
        layout.addLayout(top_layout)
        
        params_layout = QHBoxLayout()
        self.col_combo = QComboBox(); self.col_combo.setEditable(True); self.col_combo.setPlaceholderText("Col (e.g. A)")
        self.col_combo.addItems(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"])
        self.col_var_combo = QComboBox(); self.col_var_combo.addItems(self.global_variables)
        params_layout.addWidget(QLabel("Col:")); params_layout.addWidget(self.col_combo, 1); params_layout.addWidget(self.col_var_combo)
        
        self.row_label = QLabel("Row:"); self.row_spin = QSpinBox(); self.row_spin.setRange(1, 1048576)
        self.row_var_combo = QComboBox(); self.row_var_combo.addItems(self.global_variables)
        params_layout.addWidget(self.row_label); params_layout.addWidget(self.row_spin); params_layout.addWidget(self.row_var_combo)
        layout.addLayout(params_layout)
        
        assign_layout = QHBoxLayout()
        self.assign_check = QCheckBox("Assign to Var:"); self.assign_check.setChecked(True)
        self.new_var_radio = QRadioButton("New:"); self.new_var_input = QLineEdit(f"excel_res_{index+1}")
        self.existing_var_radio = QRadioButton("Ext:"); self.existing_var_combo = QComboBox(); self.existing_var_combo.addItems(["-- Select --"] + global_variables)
        assign_layout.addWidget(self.assign_check); assign_layout.addWidget(self.new_var_radio); assign_layout.addWidget(self.new_var_input)
        assign_layout.addWidget(self.existing_var_radio); assign_layout.addWidget(self.existing_var_combo)
        layout.addLayout(assign_layout)
        
        self.action_combo.currentTextChanged.connect(self._on_action_changed)
        self.col_var_combo.currentTextChanged.connect(self._on_col_var_changed)
        self.col_combo.currentTextChanged.connect(self._on_col_combo_changed)
        self.row_var_combo.currentTextChanged.connect(lambda t: self.row_spin.setDisabled(t != "-- Select Variable --"))
        self.new_var_radio.setChecked(True); self.existing_var_radio.toggled.connect(lambda c: self.existing_var_combo.setEnabled(c)); self.existing_var_combo.setEnabled(False)
        
    def _on_action_changed(self, action: str):
        is_cell = (action == "Read Cell")
        self.row_label.setVisible(is_cell); self.row_spin.setVisible(is_cell); self.row_var_combo.setVisible(is_cell)

    def _on_col_var_changed(self, text):
        is_var = (text != "-- Select Variable --")
        self.col_combo.setDisabled(is_var)

    def _on_col_combo_changed(self, text):
        if " - " in text:
            col_letter = text.split(" - ")[0]
            self.col_combo.blockSignals(True)
            self.col_combo.setEditText(col_letter)
            self.col_combo.blockSignals(False)

    def set_headers(self, headers: List[str]):
        if not headers: return
        self.col_combo.blockSignals(True)
        current = self.col_combo.currentText()
        self.col_combo.clear()
        self.col_combo.addItems(headers)
        if current: self.col_combo.setEditText(current)
        self.col_combo.blockSignals(False)

    def get_data(self) -> Dict[str, Any]:
        return {
            "type": self.action_combo.currentText(),
            "col": self.col_combo.currentText() if self.col_var_combo.currentText() == "-- Select Variable --" else "",
            "col_var": self.col_var_combo.currentText() if self.col_var_combo.currentText() != "-- Select Variable --" else "",
            "row": self.row_spin.value() if self.row_var_combo.currentText() == "-- Select Variable --" else 0,
            "row_var": self.row_var_combo.currentText() if self.row_var_combo.currentText() != "-- Select Variable --" else "",
            "assign": self.assign_check.isChecked(), "is_new_var": self.new_var_radio.isChecked(),
            "var_name": self.new_var_input.text() if self.new_var_radio.isChecked() else self.existing_var_combo.currentText()
        }

    def set_data(self, data: Dict[str, Any]):
        self.action_combo.setCurrentText(data.get("type", "Read Cell"))
        if data.get("col_var"): self.col_var_combo.setCurrentText(data["col_var"])
        else: self.col_combo.setEditText(data.get("col", ""))
        if data.get("row_var"): self.row_var_combo.setCurrentText(data["row_var"])
        else: self.row_spin.setValue(data.get("row", 1))
        self.assign_check.setChecked(data.get("assign", True))
        if data.get("is_new_var", True): self.new_var_radio.setChecked(True); self.new_var_input.setText(data.get("var_name", ""))
        else: self.existing_var_radio.setChecked(True); self.existing_var_combo.setCurrentText(data.get("var_name", "-- Select --"))

#
# --- HELPER: Action Row for Excel Write ---
#
class _ExcelWriteActionRow(QFrame):
    def __init__(self, global_variables: List[str], index: int, parent=None):
        super().__init__(parent)
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.index = index
        self.global_variables = ["-- Select Variable --"] + global_variables
        
        layout = QVBoxLayout(self)
        top_layout = QHBoxLayout()
        self.delete_btn = QPushButton("Remove"); self.delete_btn.setStyleSheet("background-color: #ff4d4d; color: white;")
        top_layout.addWidget(QLabel(f"Write Action {index+1}:")); top_layout.addStretch(); top_layout.addWidget(self.delete_btn)
        layout.addLayout(top_layout)
        
        col_layout = QHBoxLayout()
        self.col_combo = QComboBox(); self.col_combo.setEditable(True); self.col_combo.setPlaceholderText("A")
        self.col_combo.addItems(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"])
        self.col_var_combo = QComboBox(); self.col_var_combo.addItems(self.global_variables)
        col_layout.addWidget(QLabel("Col:")); col_layout.addWidget(self.col_combo, 1); col_layout.addWidget(self.col_var_combo)
        layout.addLayout(col_layout)
        
        rv_layout = QHBoxLayout()
        self.row_spin = QSpinBox(); self.row_spin.setRange(1, 1048576)
        self.row_var_combo = QComboBox(); self.row_var_combo.addItems(self.global_variables)
        self.val_edit = QLineEdit(); self.val_edit.setPlaceholderText("Value to write")
        self.val_var_combo = QComboBox(); self.val_var_combo.addItems(self.global_variables)
        rv_layout.addWidget(QLabel("Row:")); rv_layout.addWidget(self.row_spin); rv_layout.addWidget(self.row_var_combo)
        rv_layout.addWidget(QLabel("Value:")); rv_layout.addWidget(self.val_edit); rv_layout.addWidget(self.val_var_combo)
        layout.addLayout(rv_layout)
        
        self.col_var_combo.currentTextChanged.connect(self._on_col_var_changed)
        self.col_combo.currentTextChanged.connect(self._on_col_combo_changed)
        self.row_var_combo.currentTextChanged.connect(lambda t: self.row_spin.setDisabled(t != "-- Select Variable --"))
        self.val_var_combo.currentTextChanged.connect(lambda t: self.val_edit.setDisabled(t != "-- Select Variable --"))

    def _on_col_var_changed(self, text):
        is_var = (text != "-- Select Variable --")
        self.col_combo.setDisabled(is_var)
    
    def _on_col_combo_changed(self, text):
        if " - " in text:
            col_letter = text.split(" - ")[0]
            self.col_combo.blockSignals(True)
            self.col_combo.setEditText(col_letter)
            self.col_combo.blockSignals(False)

    def set_headers(self, headers: List[str]):
        if not headers: return
        self.col_combo.blockSignals(True)
        current = self.col_combo.currentText()
        self.col_combo.clear()
        self.col_combo.addItems(headers)
        if current: self.col_combo.setEditText(current)
        self.col_combo.blockSignals(False)

    def get_data(self) -> Dict[str, Any]:
        return {
            "col": self.col_combo.currentText() if self.col_var_combo.currentText() == "-- Select Variable --" else "",
            "col_var": self.col_var_combo.currentText() if self.col_var_combo.currentText() != "-- Select Variable --" else "",
            "row": self.row_spin.value() if self.row_var_combo.currentText() == "-- Select Variable --" else 0,
            "row_var": self.row_var_combo.currentText() if self.row_var_combo.currentText() != "-- Select Variable --" else "",
            "val": self.val_edit.text() if self.val_var_combo.currentText() == "-- Select Variable --" else "",
            "val_var": self.val_var_combo.currentText() if self.val_var_combo.currentText() != "-- Select Variable --" else ""
        }

    def set_data(self, data: Dict[str, Any]):
        if data.get("col_var"): self.col_var_combo.setCurrentText(data["col_var"])
        else: self.col_combo.setEditText(data.get("col", ""))
        if data.get("row_var"): self.row_var_combo.setCurrentText(data["row_var"])
        else: self.row_spin.setValue(data.get("row", 1))
        if data.get("val_var"): self.val_var_combo.setCurrentText(data["val_var"])
        else: self.val_edit.setText(data.get("val", ""))

#
# --- DIALOG: Excel Read ---
#
class _ExcelReadDialog(QDialog):
    def __init__(self, global_variables: List[str], parent: Optional[QWidget] = None, initial_config: Optional[Dict[str, Any]] = None, **kwargs):
        super().__init__(parent)
        self.setWindowTitle("Excel Read Module"); self.setMinimumSize(800, 700)
        self.global_variables = global_variables; self.current_headers = []
        
        main_layout = QVBoxLayout(self)
        file_group = QGroupBox("Excel File Selection"); file_layout = QFormLayout(file_group)
        self.file_path_edit = QLineEdit(); browse_btn = QPushButton("Browse...")
        self.file_var_combo = QComboBox(); self.file_var_combo.addItems(["-- Select Variable --"] + global_variables)
        p_row = QHBoxLayout(); p_row.addWidget(self.file_path_edit); p_row.addWidget(browse_btn)
        file_layout.addRow("File Path:", p_row); file_layout.addRow("or from Variable:", self.file_var_combo); main_layout.addWidget(file_group)
        
        sheet_group = QGroupBox("Sheet Selection"); sheet_layout = QFormLayout(sheet_group)
        self.sheet_name_edit = QLineEdit(); self.sheet_var_combo = QComboBox(); self.sheet_var_combo.addItems(["-- Select Variable --"] + global_variables)
        self.get_sheets_btn = QPushButton("Read Sheet Names & Headers")
        self.sheet_dropdown = QComboBox(); self.sheet_dropdown.addItems(["-- Select Sheet --"])
        sheet_layout.addRow("Sheet Name:", self.sheet_name_edit); sheet_layout.addRow("or from Variable:", self.sheet_var_combo)
        sheet_layout.addRow(self.get_sheets_btn); sheet_layout.addRow("Available Sheets:", self.sheet_dropdown); main_layout.addWidget(sheet_group)
        
        worksheet_group = QGroupBox("Full Worksheet Read"); worksheet_layout = QVBoxLayout(worksheet_group)
        self.read_all_check = QCheckBox("Read entire worksheet into a variable"); self.ws_var_input = QLineEdit("excel_dataframe")
        worksheet_layout.addWidget(self.read_all_check); ws_assign_layout = QHBoxLayout(); ws_assign_layout.addWidget(QLabel("Variable Name:")); ws_assign_layout.addWidget(self.ws_var_input)
        worksheet_layout.addLayout(ws_assign_layout); main_layout.addWidget(worksheet_group)
        
        actions_group = QGroupBox("Granular Actions (Max 10)"); self.actions_layout = QVBoxLayout(actions_group)
        self.add_action_btn = QPushButton("+ Add Action (Max 10)"); self.actions_scroll = QScrollArea(); self.actions_scroll.setWidgetResizable(True)
        self.actions_container = QWidget(); self.actions_list_layout = QVBoxLayout(self.actions_container); self.actions_scroll.setWidget(self.actions_container)
        self.actions_layout.addWidget(self.add_action_btn); self.actions_layout.addWidget(self.actions_scroll); main_layout.addWidget(actions_group, 1)
        
        self.action_rows: List[_ExcelReadActionRow] = []
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel); main_layout.addWidget(self.button_box)
        
        browse_btn.clicked.connect(self._browse_file); self.get_sheets_btn.clicked.connect(lambda: self._fetch_sheets_and_headers(True))
        self.sheet_dropdown.currentTextChanged.connect(self._on_sheet_selected)
        self.add_action_btn.clicked.connect(self._add_action_row)
        self.file_var_combo.currentTextChanged.connect(lambda t: self.file_path_edit.setDisabled(t != "-- Select Variable --"))
        self.sheet_var_combo.currentTextChanged.connect(lambda t: self.sheet_name_edit.setDisabled(t != "-- Select Variable --"))
        self.button_box.accepted.connect(self.accept); self.button_box.rejected.connect(self.reject)
        if initial_config: self._populate_from_config(initial_config)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel", "", "Excel Files (*.xlsx *.xls *.xlsm)")
        if path: self.file_path_edit.setText(path)

    def _fetch_sheets_and_headers(self, refresh_sheets=True):
        path = self.file_path_edit.text()
        if not path or not os.path.exists(path): QMessageBox.warning(self, "File Error", "Provide valid path."); return
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if refresh_sheets:
                self.sheet_dropdown.blockSignals(True)
                current_sel = self.sheet_dropdown.currentText()
                self.sheet_dropdown.clear(); self.sheet_dropdown.addItems(["-- Select Sheet --"] + wb.sheetnames)
                if current_sel in wb.sheetnames: self.sheet_dropdown.setCurrentText(current_sel)
                self.sheet_dropdown.blockSignals(False)
            
            sheet_name = self.sheet_name_edit.text()
            if sheet_name in wb.sheetnames:
                headers = []
                sheet = wb[sheet_name]
                for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
                    if cell.value: headers.append(f"{cell.column_letter} - {cell.value}")
                    else: headers.append(f"{cell.column_letter} - (Empty)")
                self.current_headers = headers
                for row in self.action_rows: row.set_headers(headers)
            wb.close()
        except Exception as e: QMessageBox.critical(self, "Error", str(e))

    def _on_sheet_selected(self, sheet: str):
        if sheet != "-- Select Sheet --":
            self.sheet_name_edit.setText(sheet)
            self._fetch_sheets_and_headers(refresh_sheets=False)

    def _add_action_row(self):
        if len(self.action_rows) >= 10: return
        row = _ExcelReadActionRow(self.global_variables, len(self.action_rows)); row.delete_btn.clicked.connect(lambda: self._remove_action_row(row))
        if self.current_headers: row.set_headers(self.current_headers)
        self.actions_list_layout.addWidget(row); self.action_rows.append(row)

    def _remove_action_row(self, row):
        self.actions_list_layout.removeWidget(row); self.action_rows.remove(row); row.deleteLater()
        for i, r in enumerate(self.action_rows): r.index = i

    def _populate_from_config(self, config):
        if config.get("file_var"): self.file_var_combo.setCurrentText(config["file_var"])
        else: self.file_path_edit.setText(config.get("file_path", ""))
        if config.get("sheet_var"): self.sheet_var_combo.setCurrentText(config["sheet_var"])
        else: self.sheet_name_edit.setText(config.get("sheet_name", ""))
        self.read_all_check.setChecked(config.get("read_all", False)); self.ws_var_input.setText(config.get("ws_var", "excel_dataframe"))
        for a_data in config.get("actions", []): self._add_action_row(); self.action_rows[-1].set_data(a_data)

    def get_config_data(self) -> Optional[Dict[str, Any]]:
        return {
            "file_path": self.file_path_edit.text(), "file_var": "" if self.file_var_combo.currentText() == "-- Select Variable --" else self.file_var_combo.currentText(),
            "sheet_name": self.sheet_name_edit.text(), "sheet_var": "" if self.sheet_var_combo.currentText() == "-- Select Variable --" else self.sheet_var_combo.currentText(),
            "read_all": self.read_all_check.isChecked(), "ws_var": self.ws_var_input.text(), "actions": [row.get_data() for row in self.action_rows]
        }
    def get_executor_method_name(self) -> str: return "_read_excel_action"
    def get_assignment_variable(self) -> Optional[str]: return None

#
# --- PUBLIC CLASS: Excel Read ---
#
class Excel_Read:
    def __init__(self, context: Optional[ExecutionContext] = None): self.context = context
    def _log(self, m: str): (self.context.add_log(m) if self.context else print(m))
    def configure_data_hub(self, parent_window: QWidget, global_variables: List[str], **kwargs) -> QDialog: return _ExcelReadDialog(global_variables, parent_window, **kwargs)

    def _read_excel_action(self, context: ExecutionContext, config_data: dict):
        self.context = context
        file_path = context.get_variable(config_data["file_var"]) if config_data["file_var"] else config_data["file_path"]
        sheet_name = context.get_variable(config_data["sheet_var"]) if config_data["sheet_var"] else config_data["sheet_name"]
        if not file_path or not os.path.exists(str(file_path)): raise FileNotFoundError(f"Excel file not found: {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        if config_data.get("read_all"): context.set_variable(config_data["ws_var"], pd.read_excel(file_path, sheet_name=sheet.title))
        for action in config_data.get("actions", []):
            col = context.get_variable(action["col_var"]) if action["col_var"] else action["col"]
            row_idx = context.get_variable(action["row_var"]) if action["row_var"] else action["row"]
            res = None
            if action["type"] == "Read Cell": res = sheet[f"{col}{row_idx}"].value
            elif action["type"] == "Get Last Empty Row":
                last_row = 1
                for r in range(sheet.max_row, 0, -1):
                    if sheet[f"{col}{r}"].value is not None: last_row = r; break
                res = last_row + 1
            elif action["type"] == "Get Total Line Count":
                count = 0
                for r in range(1, sheet.max_row + 1):
                    if sheet[f"{col}{r}"].value is not None: count += 1
                res = count
            elif action["type"] == "Find First Empty Row (1-Max)":
                res = sheet.max_row + 1
                for r in range(1, sheet.max_row + 1):
                    val = sheet[f"{col}{r}"].value
                    if val is None or str(val).strip() == "": res = r; break
            if action["assign"] and action["var_name"]: context.set_variable(action["var_name"], res)
        wb.close()

#
# --- DIALOG: Excel Write ---
#
class _ExcelWriteDialog(QDialog):
    def __init__(self, global_variables: List[str], parent: Optional[QWidget] = None, initial_config: Optional[Dict[str, Any]] = None, **kwargs):
        super().__init__(parent)
        self.setWindowTitle("Excel Write Module"); self.setMinimumSize(800, 700)
        self.global_variables = global_variables; self.current_headers = []
        
        main_layout = QVBoxLayout(self)
        fs_group = QGroupBox("File & Sheet Selection"); fs_layout = QFormLayout(fs_group)
        self.file_path_edit = QLineEdit(); browse_btn = QPushButton("Browse...")
        self.file_var_combo = QComboBox(); self.file_var_combo.addItems(["-- Select Variable --"] + global_variables)
        self.sheet_name_edit = QLineEdit("Sheet1"); self.sheet_var_combo = QComboBox(); self.sheet_var_combo.addItems(["-- Select Variable --"] + global_variables)
        self.get_sheets_btn = QPushButton("Read Sheet Names & Headers")
        self.sheet_dropdown = QComboBox(); self.sheet_dropdown.addItems(["-- Select Sheet --"])
        
        p_row = QHBoxLayout(); p_row.addWidget(self.file_path_edit); p_row.addWidget(browse_btn)
        fs_layout.addRow("File Path:", p_row); fs_layout.addRow("File Variable:", self.file_var_combo)
        fs_layout.addRow("Sheet Name:", self.sheet_name_edit); fs_layout.addRow("Sheet Variable:", self.sheet_var_combo)
        fs_layout.addRow(self.get_sheets_btn); fs_layout.addRow("Available Sheets:", self.sheet_dropdown); main_layout.addWidget(fs_group)
        
        actions_group = QGroupBox("Write Actions (Max 10)"); self.actions_layout = QVBoxLayout(actions_group)
        self.add_action_btn = QPushButton("+ Add Write Action (Max 10)"); self.actions_scroll = QScrollArea(); self.actions_scroll.setWidgetResizable(True)
        self.actions_container = QWidget(); self.actions_list_layout = QVBoxLayout(self.actions_container); self.actions_scroll.setWidget(self.actions_container)
        self.actions_layout.addWidget(self.add_action_btn); self.actions_layout.addWidget(self.actions_scroll); main_layout.addWidget(actions_group, 1)
        
        self.action_rows: List[_ExcelWriteActionRow] = []
        self.save_check = QCheckBox("Save workbook after writing"); self.save_check.setChecked(True); main_layout.addWidget(self.save_check)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel); main_layout.addWidget(self.button_box)
        
        browse_btn.clicked.connect(self._browse_file); self.get_sheets_btn.clicked.connect(lambda: self._fetch_sheets_and_headers(True))
        self.sheet_dropdown.currentTextChanged.connect(self._on_sheet_selected)
        self.add_action_btn.clicked.connect(self._add_action_row)
        self.file_var_combo.currentTextChanged.connect(lambda t: self.file_path_edit.setDisabled(t != "-- Select Variable --"))
        self.sheet_var_combo.currentTextChanged.connect(lambda t: self.sheet_name_edit.setDisabled(t != "-- Select Variable --"))
        self.button_box.accepted.connect(self.accept); self.button_box.rejected.connect(self.reject)
        if initial_config: self._populate_from_config(initial_config)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel", "", "Excel Files (*.xlsx *.xls)")
        if path: self.file_path_edit.setText(path)

    def _fetch_sheets_and_headers(self, refresh_sheets=True):
        path = self.file_path_edit.text()
        if not path or not os.path.exists(path): QMessageBox.warning(self, "File Error", "Provide valid path."); return
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if refresh_sheets:
                self.sheet_dropdown.blockSignals(True)
                current_sel = self.sheet_dropdown.currentText()
                self.sheet_dropdown.clear(); self.sheet_dropdown.addItems(["-- Select Sheet --"] + wb.sheetnames)
                if current_sel in wb.sheetnames: self.sheet_dropdown.setCurrentText(current_sel)
                self.sheet_dropdown.blockSignals(False)
            
            sheet_name = self.sheet_name_edit.text()
            if sheet_name in wb.sheetnames:
                headers = []
                sheet = wb[sheet_name]
                for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
                    if cell.value: headers.append(f"{cell.column_letter} - {cell.value}")
                    else: headers.append(f"{cell.column_letter} - (Empty)")
                self.current_headers = headers
                for row in self.action_rows: row.set_headers(headers)
            wb.close()
        except Exception as e: QMessageBox.critical(self, "Error", str(e))

    def _on_sheet_selected(self, sheet_name):
        if sheet_name != "-- Select Sheet --":
            self.sheet_name_edit.setText(sheet_name)
            self._fetch_sheets_and_headers(refresh_sheets=False)

    def _add_action_row(self):
        if len(self.action_rows) >= 10: return
        row = _ExcelWriteActionRow(self.global_variables, len(self.action_rows)); row.delete_btn.clicked.connect(lambda: self._remove_action_row(row))
        if self.current_headers: row.set_headers(self.current_headers)
        self.actions_list_layout.addWidget(row); self.action_rows.append(row)

    def _remove_action_row(self, row):
        self.actions_list_layout.removeWidget(row); self.action_rows.remove(row); row.deleteLater()
        for i, r in enumerate(self.action_rows): r.index = i

    def _populate_from_config(self, config):
        self.file_path_edit.setText(config.get("file_path", "")); self.file_var_combo.setCurrentText(config.get("file_var", "-- Select Variable --"))
        self.sheet_name_edit.setText(config.get("sheet_name", "Sheet1")); self.sheet_var_combo.setCurrentText(config.get("sheet_var", "-- Select Variable --"))
        for a_data in config.get("actions", []): self._add_action_row(); self.action_rows[-1].set_data(a_data)
        self.save_check.setChecked(config.get("save", True))

    def get_config_data(self) -> Optional[Dict[str, Any]]:
        return {
            "file_path": self.file_path_edit.text(), "file_var": "" if self.file_var_combo.currentText() == "-- Select Variable --" else self.file_var_combo.currentText(),
            "sheet_name": self.sheet_name_edit.text(), "sheet_var": "" if self.sheet_var_combo.currentText() == "-- Select Variable --" else self.sheet_var_combo.currentText(),
            "actions": [row.get_data() for row in self.action_rows], "save": self.save_check.isChecked()
        }
    def get_executor_method_name(self) -> str: return "_write_excel_action"
    def get_assignment_variable(self) -> Optional[str]: return None

#
# --- PUBLIC CLASS: Excel Write ---
#
class Excel_Write:
    def __init__(self, context: Optional[ExecutionContext] = None): self.context = context
    def _log(self, m: str): (self.context.add_log(m) if self.context else print(m))
    def configure_data_hub(self, parent_window: QWidget, global_variables: List[str], **kwargs) -> QDialog: return _ExcelWriteDialog(global_variables, parent_window, **kwargs)

    def _write_excel_action(self, context: ExecutionContext, config_data: dict):
        self.context = context
        file_path = context.get_variable(config_data["file_var"]) if config_data["file_var"] else config_data["file_path"]
        sheet_name = context.get_variable(config_data["sheet_var"]) if config_data["sheet_var"] else config_data["sheet_name"]
        if not file_path or not os.path.exists(str(file_path)): raise FileNotFoundError(f"Excel file not found: {file_path}")
        wb = openpyxl.load_workbook(file_path); sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        for action in config_data.get("actions", []):
            col = context.get_variable(action["col_var"]) if action["col_var"] else action["col"]
            row_idx = context.get_variable(action["row_var"]) if action["row_var"] else action["row"]
            val = context.get_variable(action["val_var"]) if action["val_var"] else action["val"]
            sheet[f"{col}{row_idx}"] = val
            self._log(f"Wrote '{val}' to {col}{row_idx}")
        if config_data.get("save"): wb.save(file_path)
        wb.close()
