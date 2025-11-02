import sys
import json
import win32com.client
import win32gui
import os
import clipboard
from datetime import datetime, timedelta, date
import pandas as pd
import numpy as np
import time
import io
import base64
from PIL import ImageGrab
import PIL.Image
import sys
import json
import subprocess
import pyautogui
import re
from PIL.ImageQt import ImageQt
from pymsgbox import *
import pygetwindow as gw
import keyboard
import openpyxl
import datetime
from datetime import datetime
from my_lib.shared_context import ExecutionContext as Context
from typing import List, Dict, Any, Optional, Union
class handle_excel():
    """A class for handling Excel file operations using the openpyxl library."""

    def __init__(self,context: Context):
        """Initializes the File_handle class."""
        self.context = context
        pass
    def read_excel(self,file_link,sheet_name=None):
        """Opens an Excel file and returns the workbook and a specific sheet object.

        Args:
            file_link (str): The full path to the Excel file.
            sheet_name (str, optional): The name of the sheet to access. If None or not
                                        found, the active sheet is returned. Defaults to None.

        Returns:
            list or None: A list containing [workbook, sheet] objects on success,
                          or None if an error occurs (e.g., file not found).
        """
        print (file_link)
        try:
            workbook = openpyxl.load_workbook(file_link)
            sheet = workbook.active
            if sheet_name is not None and sheet_name !='None'  and sheet_name !='':
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    return [workbook, sheet]
                else:
                    print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
                    sheet = workbook.active
                    return [workbook, sheet]
            else:
                sheet = workbook.active  # Get the active (first) sheet
                return [workbook, sheet]
    
    
        except FileNotFoundError:
            print(f"Error: File not found at '{file_link}'")
            return None
        except Exception as e:
            print(f"An error occurred while reading the Excel file: {e}")
            return None
            
        return None
    def check_Length_String (self,String):
        """ Check lenght of a String """
        try:
            str_len = len(str(String))
        except:
            str_len = 0
        return str_len

    def read_excel_cells(self,workbook,col,row : int):
        """Reads the value from a single cell in an Excel worksheet.

        Args:
            workbook (list): The [workbook, sheet] list object returned by `read_excel`.
            col (str): The column letter of the cell (e.g., 'A', 'B').
            row (int): The 0-indexed row number. The method adds 1 to match Excel's
                       1-based row indexing.

        Returns:
            any: The value contained in the specified cell.
        """
        
        col=col.replace("'","")
        print (f'"{col}{int(row)}"')
        sheet = workbook[1]
        
        return sheet[f"{col}{int(row)}"].value
        
    def write_excel_cells(self,workbook,col,row,value):
        """Writes a value to a single cell in an Excel worksheet.

        Args:
            workbook (list): The [workbook, sheet] list object returned by `read_excel`.
            col (str): The column letter of the cell.
            row (int): The 0-indexed row number. The method adds 1 to match Excel's
                       1-based row indexing.
            value (any): The value to write into the cell.

        Returns:
            str: A confirmation message: "assigned Value".
        """
        workbook[1][f"{col}{int(row+1)}"] = value
        return "assigned Value"
        
    def save_excel(self,workbook,file_name):
        """Saves and closes the Excel workbook to a specified file.

        Args:
            workbook (list): The [workbook, sheet] list object to be saved.
            file_name (str): The file path (including name) to save the workbook to.

        Returns:
            str: A confirmation message: "saved excel".
        """
        workbook[0].save(file_name)
        workbook[0].close()
        return f"saved excel"  
        
    def close_excel(self,workbook):
        """Closes the Excel workbook without saving changes.

        Args:
            workbook (openpyxl.workbook.workbook.Workbook): The workbook object to close.

        Returns:
            str: A confirmation message: "closed excel".
        """
        workbook[0].close()
        return f"closed excel" 
    
    def count_non_empty_rows_in_column(self,workbook, column_identifier):
        """Counts the number of non-empty cells in a specific column of a worksheet.

        Args:
            workbook (list): The [workbook, sheet] list object.
            column_identifier (str or int): The column to check. Can be a column letter
                                            (e.g., 'A') or a 1-based integer index (e.g., 1).

        Returns:
            int: The number of non-empty cells in the specified column, or -1 if an error occurs.
        """
        worksheet_object =workbook[1]
        if not isinstance(worksheet_object, openpyxl.worksheet.worksheet.Worksheet):
            print("Error: The provided object is not an openpyxl Worksheet.")
            return -1
    
        count = 0
        try:
            # Get the max row of the entire sheet to avoid iterating endlessly
            max_sheet_row = worksheet_object.max_row
    
            # Convert column identifier to column letter if it's an integer
            if isinstance(column_identifier, int):
                if column_identifier <= 0:
                    print("Error: Column index must be 1-based.")
                    return -1
                column_letter = openpyxl.utils.get_column_letter(column_identifier)
            elif isinstance(column_identifier, str):
                column_letter = column_identifier.upper() # Ensure uppercase
                # Optional: Validate column letter
                if not column_letter.isalpha():
                    print(f"Error: Invalid column letter '{column_identifier}'.")
                    return -1
            else:
                print("Error: Column identifier must be a string (letter) or an integer (1-based index).")
                return -1
    
            # Iterate through cells in the specified column up to the max row of the sheet
            for row_num in range(1, max_sheet_row + 1):
                cell = worksheet_object[f"{column_letter}{row_num}"]
                if cell.value is not None:
                    count += 1
    
            print(f"Sheet '{worksheet_object.title}', Column '{column_letter}': Non-empty cells = {count}")
            return count
    
        except Exception as e:
            print(f"An error occurred while counting column '{column_identifier}': {e}")
            return -1
